# -*- coding: utf-8 -*-
"""
Convertitore: maxi-impianti.xlsx  ->  js/maxi-data.js

Legge il file Excel con gli impianti maxi e rigenera il file dati che
alimenta SIA la galleria SIA la mappa della pagina maxi.html.

USO: doppio-click su AGGIORNA_MAPPA_MAXI.bat  (oppure:  python build-maxi-data.py)

Colonne attese (prima riga = intestazioni, in qualsiasi ordine):
  code, city, pos, type, dim, sqm, light, flow, photo, lat, lng
"""
import json, sys, os
import openpyxl

XLSX = 'maxi-impianti.xlsx'
OUT  = os.path.join('js', 'maxi-data.js')

def truthy(v):
    if v is None:
        return False
    return str(v).strip().lower() in ('si', 'sì', 's', 'yes', 'y', 'true', '1', 'x')

def num(v):
    if v is None or str(v).strip() == '':
        return None
    try:
        return float(str(v).replace(',', '.'))
    except ValueError:
        return None

def main():
    if not os.path.exists(XLSX):
        print('ERRORE: non trovo il file %s in questa cartella.' % XLSX)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['Impianti'] if 'Impianti' in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print('ERRORE: il foglio è vuoto.'); sys.exit(1)

    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    required = ['code', 'city', 'pos', 'type', 'dim']
    missing = [c for c in required if c not in idx]
    if missing:
        print('ERRORE: mancano le colonne: %s' % ', '.join(missing)); sys.exit(1)

    def cell(r, name):
        i = idx.get(name)
        if i is None or i >= len(r):
            return None
        return r[i]

    impianti = []
    for r in rows[1:]:
        code = cell(r, 'code')
        if code is None or str(code).strip() == '':
            continue  # salta righe vuote
        if str(code).strip().lower() in ('code', 'codice'):
            continue  # salta l'eventuale riga di etichette leggibili
        sqm_v = num(cell(r, 'sqm'))
        lat = num(cell(r, 'lat'))
        lng = num(cell(r, 'lng'))
        item = {
            'code':  str(code).strip(),
            'city':  str(cell(r, 'city')  or '').strip(),
            'pos':   str(cell(r, 'pos')   or '').strip(),
            'type':  str(cell(r, 'type')  or '').strip(),
            'dim':   str(cell(r, 'dim')   or '').strip(),
            'sqm':   int(sqm_v) if sqm_v is not None else None,
            'light': truthy(cell(r, 'light')),
            'flow':  str(cell(r, 'flow')  or '').strip(),
            'photo': str(cell(r, 'photo') or '').strip(),
        }
        if lat is not None and lng is not None:
            item['lat'] = lat
            item['lng'] = lng
        impianti.append(item)

    body = json.dumps(impianti, ensure_ascii=False, indent=2)
    js = ('/* File generato automaticamente da build-maxi-data.py — NON modificare a mano.\n'
          '   Per aggiornare: modifica maxi-impianti.xlsx e ri-lancia AGGIORNA_MAPPA_MAXI.bat */\n'
          'window.MAXI_IMPIANTI = ' + body + ';\n')

    with open(OUT, 'w', encoding='utf-8', newline='\n') as f:
        f.write(js)

    n_map = sum(1 for i in impianti if 'lat' in i)
    print('OK: %d impianti scritti in %s (%d con coordinate sulla mappa).' % (len(impianti), OUT, n_map))
    if n_map < len(impianti):
        print('Nota: %d impianti senza lat/lng non compariranno sulla mappa.' % (len(impianti) - n_map))

if __name__ == '__main__':
    main()
